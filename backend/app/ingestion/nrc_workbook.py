from collections import defaultdict
from datetime import date, datetime
from io import BytesIO
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.ingestion.delimited_upload import as_float, as_int
from app.ingestion.unit_conversion import to_gallons


NRC_REQUIRED_SHEETS = {"INCIDENT_COMMONS", "INCIDENT_DETAILS", "MATERIAL_INVOLVED"}


def _json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def _rows(workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
    sheet = workbook[sheet_name]
    values = sheet.iter_rows(values_only=True)
    headers = [str(value) for value in next(values)]
    return [
        {header: _json_value(value) for header, value in zip(headers, row, strict=False)}
        for row in values
    ]


def parse_nrc_workbook(content: bytes) -> tuple[list[dict[str, Any]], dict[str, list[str]]]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError) as exc:
        raise ValueError("NRC upload is not a readable XLSX workbook.") from exc
    missing_sheets = NRC_REQUIRED_SHEETS.difference(workbook.sheetnames)
    if missing_sheets:
        raise ValueError(f"NRC workbook is missing required sheet(s): {', '.join(sorted(missing_sheets))}.")
    common_rows = _rows(workbook, "INCIDENT_COMMONS")
    detail_rows = _rows(workbook, "INCIDENT_DETAILS")
    material_rows = _rows(workbook, "MATERIAL_INVOLVED")
    headers = {
        sheet_name: [str(cell.value) for cell in workbook[sheet_name][1]]
        for sheet_name in sorted(NRC_REQUIRED_SHEETS)
    }
    return join_nrc_records(common_rows, detail_rows, material_rows), headers


def join_nrc_records(
    common_rows: list[dict[str, Any]],
    detail_rows: list[dict[str, Any]],
    material_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    common_by_id = {str(row["SEQNOS"]): row for row in common_rows if row.get("SEQNOS") is not None}
    detail_by_id = {str(row["SEQNOS"]): row for row in detail_rows if row.get("SEQNOS") is not None}
    materials_by_id: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in material_rows:
        if row.get("SEQNOS") is not None:
            materials_by_id[str(row["SEQNOS"])].append(row)
    records: list[dict[str, Any]] = []
    for report_id, common in common_by_id.items():
        details = detail_by_id.get(report_id, {})
        materials = materials_by_id.get(report_id, [])
        names = sorted(
            {
                str(material.get("NAME_OF_MATERIAL")).strip()
                for material in materials
                if material.get("NAME_OF_MATERIAL")
            }
        )
        converted = [
            to_gallons(as_float(material.get("AMOUNT_OF_MATERIAL")), material.get("UNIT_OF_MEASURE"))
            for material in materials
        ]
        converted = [item for item in converted if item]
        unconverted_units = sorted(
            {
                str(material.get("UNIT_OF_MEASURE")).strip()
                for material in materials
                if material.get("UNIT_OF_MEASURE")
                and to_gallons(as_float(material.get("AMOUNT_OF_MATERIAL")), material.get("UNIT_OF_MEASURE"))
                is None
            }
        )
        records.append(
            {
                "report_id": report_id,
                "event_date": str(common.get("INCIDENT_DATE_TIME") or "")[:10] or None,
                "region": common.get("LOCATION_STATE"),
                "narrative": common.get("DESCRIPTION_OF_INCIDENT"),
                "commodities": names,
                "commodity": " | ".join(names)[:255] or None,
                "materials": materials,
                "severity_features": {
                    "source_system": "NRC",
                    "source_capability": "count_bearing",
                    "injuries_count": as_int(details.get("NUMBER_INJURED")),
                    "fatalities_count": as_int(details.get("NUMBER_FATALITIES")),
                    "evacuations_count": as_int(details.get("NUMBER_EVACUATED")),
                    "commodities": names,
                    "raw_material_row_count": len(materials),
                    "quantity_released_gallons": sum(item.gallons for item in converted),
                    "quantity_gallons_approximate": any(item.approximate for item in converted),
                    "converted_liquid_units": sorted({item.source_unit for item in converted}),
                    "unconverted_material_units": unconverted_units,
                },
                "source_rows": {"incident_common": common, "incident_details": details, "materials": materials},
            }
        )
    return records
