import os
import re
from io import BytesIO

os.environ["DATABASE_URL"] = "sqlite+pysqlite:///./test-cbrne.db"
os.environ["DATA_DIR"] = "./test-data"

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.database import Base, engine
from app.main import app
from app.routers import connectors


client = TestClient(app)


def nrc_fixture_workbook() -> bytes:
    workbook = Workbook()
    common = workbook.active
    common.title = "INCIDENT_COMMONS"
    common.append(["SEQNOS", "INCIDENT_DATE_TIME", "LOCATION_STATE", "DESCRIPTION_OF_INCIDENT"])
    common.append(["NRC-TEST-1", "2026-01-08 10:00", "TX", "Safe NRC fixture incident"])
    details = workbook.create_sheet("INCIDENT_DETAILS")
    details.append(["SEQNOS", "NUMBER_INJURED", "NUMBER_FATALITIES", "NUMBER_EVACUATED"])
    details.append(["NRC-TEST-1", 5, 0, 110])
    material = workbook.create_sheet("MATERIAL_INVOLVED")
    material.append(["SEQNOS", "NAME_OF_MATERIAL", "AMOUNT_OF_MATERIAL", "UNIT_OF_MEASURE"])
    material.append(["NRC-TEST-1", "CHLORINE", 12000, "GALLON(S)"])
    material.append(["NRC-TEST-1", "AMMONIA", 1, "POUND(S)"])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def setup_module() -> None:
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)


def test_health() -> None:
    response = client.get("/health")
    assert response.status_code == 200
    assert response.json()["status"] == "ok"


def test_ingest_detect_review_and_plan_audit() -> None:
    source = client.post(
        "/sources",
        json={
            "name": "Approved Hazmat Extract",
            "organization": "Test Fixture",
            "url": "https://example.test/source",
            "source_type": "PUBLIC_DATASET",
            "modality": "CHEM",
            "access_terms": "Safe fixture for automated testing.",
            "limitations": "This is a test fixture and not an operational source record.",
        },
    )
    assert source.status_code == 201
    source_id = source.json()["id"]
    fixture = (
        "incident_id,event_date,type,region,commodity,injuries,fatalities,evacuated,narrative\n"
        "TEST-001,2026-01-01,chemical release,Region A,Test Material,6,0,0,Safe test fixture\n"
        "TEST-002,2026-01-02,chemical release,Region A,Test Material,0,0,0,Safe test fixture\n"
        "TEST-003,2026-01-03,chemical release,Region A,Test Material,0,0,0,Safe test fixture\n"
    )
    upload = client.post(
        f"/ingests/upload?source_id={source_id}",
        files={"file": ("fixture.csv", fixture, "text/csv")},
    )
    assert upload.status_code == 201
    batch_id = upload.json()["id"]
    normalize = client.post(
        f"/ingests/{batch_id}/normalize",
        json={
            "version": "test-v1",
            "fields": {
                "source_record_id": "incident_id",
                "event_date": "event_date",
                "event_type": "type",
                "region": "region",
                "commodity": "commodity",
                "injuries": "injuries",
                "fatalities": "fatalities",
                "evacuated": "evacuated",
                "narrative": "narrative",
            },
            "hazard_domain": "CHEM",
        },
    )
    assert normalize.status_code == 200
    run = client.post("/detections/run", json={"ingest_batch_id": batch_id})
    assert run.status_code == 200
    assert run.json()["alerts_created"] >= 2
    alerts = client.get("/alerts").json()
    alert_id = alerts[0]["id"]
    review = client.post(
        f"/alerts/{alert_id}/reviews",
        json={
            "reviewer": "test_analyst",
            "disposition": "ESCALATE",
            "threat_level": "TL3",
            "note": "Test review only.",
        },
    )
    assert review.status_code == 200
    notification = client.post(
        f"/alerts/{alert_id}/notifications",
        json={
            "threat_level": "TL3",
            "route_type": "INTERNAL",
            "route_name": "Duty Lead",
            "reporting_assessment": "REVIEW_REQUIRED",
            "rationale": "Test notification record only.",
        },
    )
    assert notification.status_code == 201
    plan = client.post(
        f"/alerts/{alert_id}/plan-reviews",
        json={
            "plan_code": "NIMS_ICS",
            "applicability": "POTENTIALLY_APPLICABLE",
            "rationale": "Test doctrine record only.",
            "reviewer": "test_analyst",
        },
    )
    assert plan.status_code == 201
    detail = client.get(f"/alerts/{alert_id}").json()
    assert detail["confirmed_threat_level"] == "TL3"
    assert detail["notifications"][0]["reporting_assessment"] == "REVIEW_REQUIRED"
    assert detail["plan_reviews"][0]["activation_status"] == "NOT_VERIFIED"


def test_noaa_connector_normalizes_official_shape(monkeypatch) -> None:
    csv_content = (
        "id,open_date,name,location,threat,tags,commodity,max_ptl_release_gallons,description\n"
        "NOAA-TEST-1,2026-01-04,Safe connector fixture,Test Coast,Chemical,Pipeline,"
        "Chlorine Gas,100,Safe connector test only\n"
    ).encode()

    class FakeResponse:
        content = csv_content

        def raise_for_status(self) -> None:
            return None

    monkeypatch.setattr(connectors.httpx, "get", lambda *args, **kwargs: FakeResponse())
    response = client.post("/connectors/noaa-incidentnews/sync")
    assert response.status_code == 201
    assert response.json()["records_received"] == 1
    assert response.json()["chemical_events"] == 1
    events = client.get("/events?domain=CHEM").json()
    event = next(event for event in events if event["source_record_id"] == "NOAA-TEST-1")
    assert event["commodity"] == "Chlorine Gas"
    run = client.post("/detections/run", json={"ingest_batch_id": response.json()["ingest_batch_id"]})
    assert run.status_code == 200
    assert run.json()["rule_set_version"] == "CHEM_HAZMAT_V0.4"
    assert run.json()["alerts_created"] == 1
    alerts = client.get("/alerts").json()
    assert any("EPA RMP toxic substance commodity match: NOAA-TEST-1" in alert["title"] for alert in alerts)


def test_phmsa_export_maps_consequences_for_detection() -> None:
    fixture = (
        "Incident ID\tDate Of Incident\tIncident City\tIncident State\tCommodity Long Name\t"
        "Total Hazmat Fatalities\tHazmat Injury Indicator\tSerious Evacuations\tQuantity Released\tUnit Of Measure\t"
        "Description of Events\n"
        "PHMSA-TEST-1\t2026-01-05\tTest City\tMI\tTest Material\t0\tYes\tYes\t60000\tLGA\t"
        "Safe PHMSA fixture line one\n"
        "PHMSA-TEST-1\t2026-01-05\tTest City\tMI\tTest Material\t0\tYes\tYes\t60000\tLGA\t"
        "Safe PHMSA duplicate incident line\n"
        "PHMSA-TEST-GCF\t2026-01-06\tTest City\tMI\tTest Gas\t0\tNo\tNo\t1000000\tGCF\t"
        "Safe unconverted gas quantity line\n"
    )
    imported = client.post(
        "/connectors/phmsa-hazmat/import",
        files={"file": ("phmsa-fixture.txt", fixture, "text/plain")},
    )
    assert imported.status_code == 201
    assert imported.json()["mapping_version"] == "phmsa-hazmat-export-v3"
    events = client.get("/events?domain=CHEM").json()
    event = next(event for event in events if event["source_record_id"] == "PHMSA-TEST-1")
    assert event["commodity"] == "Test Material"
    assert event["event_date"] == "2026-01-05"
    assert event["severity_features"]["fatalities"] == 0
    assert event["severity_features"]["injuries"] == 1
    assert event["severity_features"]["evacuated"] == 1
    assert event["severity_features"]["quantity_released_liquid_gallons"] == 60000
    gas_event = next(event for event in events if event["source_record_id"] == "PHMSA-TEST-GCF")
    assert "quantity_released_liquid_gallons" not in gas_event["severity_features"]
    run = client.post("/detections/run", json={"ingest_batch_id": imported.json()["ingest_batch_id"]})
    assert run.status_code == 200
    assert run.json()["rule_set_version"] == "CHEM_HAZMAT_V0.4"
    assert run.json()["alerts_created"] == 2
    alerts = client.get("/alerts").json()
    consequence = next(alert for alert in alerts if alert["title"] == "Reported consequence severity: PHMSA-TEST-1")
    release = next(
        alert for alert in alerts if alert["title"] == "Large reported liquid-gallon release quantity: PHMSA-TEST-1"
    )
    assert consequence["recommended_threat_level"] == "TL2"
    assert release["score"] == 70
    assert not any("NOAA-TEST-1" in alert["title"] for alert in alerts)
    history = client.get("/alerts?include_history=true").json()
    assert any("NOAA-TEST-1" in alert["title"] for alert in history)
    summary = client.get("/metrics/summary").json()
    assert summary["open_alerts"] == 2
    assert summary["current_rule_set_version"] == "CHEM_HAZMAT_V0.4"


def test_nrc_workbook_count_rules_and_cross_source_correlation() -> None:
    phmsa_fixture = (
        "Incident ID\tDate Of Incident\tIncident City\tIncident State\tCommodity Long Name\t"
        "Total Hazmat Fatalities\tHazmat Injury Indicator\tSerious Evacuations\tQuantity Released\tUnit Of Measure\t"
        "Description of Events\n"
        "PHMSA-NRC-CORR\t2026-01-07\tFixture City\tTX\tChlorine\t0\tNo\tNo\t0\tLGA\t"
        "Safe correlation fixture\n"
    )
    phmsa = client.post(
        "/connectors/phmsa-hazmat/import",
        files={"file": ("phmsa-correlation-fixture.txt", phmsa_fixture, "text/plain")},
    )
    assert phmsa.status_code == 201

    imported = client.post(
        "/connectors/nrc/import",
        files={"file": ("nrc-fixture.xlsx", nrc_fixture_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert imported.status_code == 201
    assert imported.json()["mapping_version"] == "nrc-cy-workbook-v1"
    assert imported.json()["chemical_events"] == 1
    events = client.get("/events?domain=CHEM").json()
    event = next(event for event in events if event["source_record_id"] == "NRC-TEST-1")
    assert event["severity_features"]["injuries_count"] == 5
    assert event["severity_features"]["evacuations_count"] == 110
    assert event["severity_features"]["raw_material_row_count"] == 2
    assert event["severity_features"]["quantity_released_gallons"] == 12000

    run = client.post("/detections/run", json={"ingest_batch_id": imported.json()["ingest_batch_id"]})
    assert run.status_code == 200
    assert run.json()["rule_set_version"] == "CHEM_HAZMAT_V0.4"
    assert run.json()["alerts_created"] == 5
    alerts = client.get("/alerts").json()
    titles = {alert["title"] for alert in alerts}
    assert "High reported injury count: NRC-TEST-1" in titles
    assert "Large reported evacuation count: NRC-TEST-1" in titles
    assert "Large reported liquid-gallon release quantity: NRC-TEST-1" in titles
    assert "EPA RMP toxic substance commodity match: NRC-TEST-1" in titles
    correlated = next(
        alert for alert in alerts if alert["title"] == "Cross-source chemical incident proximity: NRC-TEST-1"
    )
    assert correlated["result_label"] == "CORRELATED_ALERT"
    assert correlated["recommended_threat_level"] == "TL2"


def test_ai_misuse_safe_fixture_routes_internal_review_only() -> None:
    imported = client.post("/ai-misuse/import-safe-evaluation")
    assert imported.status_code == 201
    assert imported.json()["records_received"] == 34
    batch_id = imported.json()["ingest_batch_id"]
    events = client.get("/events?domain=AI_MISUSE&limit=100").json()
    assert len(events) == 34
    assert all(event["hazard_domain"] == "AI_MISUSE" for event in events)
    prohibited_public_fixture_terms = {
        "quantity",
        "material",
        "concentration",
        "dosage",
        "acquire",
        "purchase",
        "procure",
        "disperse",
        "delivery method",
        "release method",
        "recipe",
        "protocol",
    }
    for event in events:
        narrative = (event["narrative"] or "").lower()
        assert not re.search(r"\d", narrative)
        assert not any(term in narrative for term in prohibited_public_fixture_terms)

    run = client.post(
        "/detections/run",
        json={"ingest_batch_id": batch_id, "domain_pack": "AI_MISUSE"},
    )
    assert run.status_code == 200
    assert run.json()["rule_set_version"] == "AI_MISUSE_V0.1"
    assert run.json()["alerts_created"] == 35
    alerts = client.get("/alerts?domain_pack=AI_MISUSE").json()
    assert all(alert["review_framework"] == "AI_MISUSE_REVIEW" for alert in alerts)
    assert any(alert["recommended_review_level"] == "MR1" for alert in alerts)
    assert any(alert["recommended_review_level"] == "MR2" for alert in alerts)
    mr3 = next(alert for alert in alerts if alert["recommended_review_level"] == "MR3")

    detail = client.get(f"/alerts/{mr3['id']}").json()
    assert detail["recommended_threat_level"] == "N/A"
    review = client.post(
        f"/alerts/{mr3['id']}/reviews",
        json={
            "reviewer": "safety_reviewer",
            "disposition": "INVESTIGATE",
            "review_level": "MR3",
            "note": "Safe fixture safety-review test.",
        },
    )
    assert review.status_code == 200
    assert review.json()["confirmed_review_level"] == "MR3"
    blocked_notification = client.post(
        f"/alerts/{mr3['id']}/notifications",
        json={
            "threat_level": "TL3",
            "route_type": "INTERNAL",
            "route_name": "Not applicable",
            "reporting_assessment": "NOT_APPLICABLE",
            "rationale": "Must be rejected for fixture alert.",
        },
    )
    assert blocked_notification.status_code == 409
    blocked_plan = client.post(
        f"/alerts/{mr3['id']}/plan-reviews",
        json={
            "plan_code": "NIMS_ICS",
            "applicability": "NOT_APPLICABLE",
            "rationale": "Must be rejected for fixture alert.",
            "reviewer": "safety_reviewer",
        },
    )
    assert blocked_plan.status_code == 409
    evaluation = client.get("/ai-misuse/evaluation/latest")
    assert evaluation.status_code == 200
    result = evaluation.json()
    assert result["cases_evaluated"] == 34
    assert result["exact_routing_agreement"] == 34
    assert result["missed_escalation_cases"] == 0
    assert result["unexpected_escalations"] == 0
    assert "Fixture conformance only" in result["claim_limit"]


def test_unknown_detection_domain_is_rejected() -> None:
    response = client.post("/detections/run", json={"domain_pack": "UNKNOWN_PACK"})
    assert response.status_code == 422
